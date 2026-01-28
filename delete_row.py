import argparse
import sys
from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(
        description="Delete a row from an Excel file."
    )

    parser.add_argument(
        "input",
        help="Path to input xlsx file"
    )

    parser.add_argument(
        "line",
        type=int,
        help="Row number to delete (1-based index)"
    )

    parser.add_argument(
        "--sheet",
        help="Sheet name (required if multiple sheets exist)",
        default=None
    )

    parser.add_argument(
        "--output",
        help="Output file (optional, overwrites input if omitted)",
        default=None
    )

    args = parser.parse_args()

    # Load workbook
    wb = load_workbook(args.input)

    # Determine sheet
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"Error: Sheet '{args.sheet}' not found.")
            sys.exit(1)
        ws = wb[args.sheet]
    else:
        if len(wb.sheetnames) > 1:
            print("Error: Multiple sheets present. Please specify --sheet.")
            print("Available sheets:", ", ".join(wb.sheetnames))
            sys.exit(1)
        ws = wb.active

    # Validate row number
    if args.line < 1 or args.line > ws.max_row:
        print(f"Error: Row {args.line} is out of range (1 - {ws.max_row}).")
        sys.exit(1)

    # Delete row
    ws.delete_rows(args.line)

    # Determine output file
    output_file = args.output if args.output else args.input

    # Save
    wb.save(output_file)

    print(f"Row {args.line} deleted successfully.")
    print(f"Saved to: {output_file}")


if __name__ == "__main__":
    main()
